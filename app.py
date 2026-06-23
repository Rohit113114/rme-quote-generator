import re
import smtplib
import time
import random
from datetime import datetime, date
from io import BytesIO
from pathlib import Path
from xml.sax.saxutils import escape
from email.message import EmailMessage

import gspread
import pandas as pd
import streamlit as st
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from pyluach import dates
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


st.set_page_config(page_title="RME Commercial Dashboard", layout="wide")


def check_login():
    USERNAME = "admin"
    PASSWORD = st.secrets["auth"]["password"]

    query_params = st.query_params

    if query_params.get("login") == "true":
        st.session_state["logged_in"] = True

    if "logged_in" not in st.session_state:
        st.session_state["logged_in"] = False

    if st.session_state["logged_in"]:
        return True

    st.title("RME Commercial Dashboard Login")

    username = st.text_input("Username")
    password = st.text_input("Password", type="password")

    if st.button("Login"):
        if username == USERNAME and password == PASSWORD:
            st.session_state["logged_in"] = True
            st.query_params["login"] = "true"
            st.rerun()
        else:
            st.error("Invalid username or password")

    return False


col1, col2 = st.columns([9, 1])

with col2:
    if st.button("Logout"):
        st.session_state["logged_in"] = False
        st.query_params.clear()
        st.rerun()


st.markdown(
    """
    <style>
        .rme-header {
            background-color: #111827;
            padding: 12px 24px;
            border-radius: 10px;
            display: flex;
            align-items: center;
            gap: 30px;
            margin-bottom: 25px;
        }

        .rme-header img {
            width: 130px;
        }

        .rme-header-title {
            color: white;
            font-size: 34px;
            font-weight: 700;
            margin: 0;
        }

        .rme-header-subtitle {
            color: #D1D5DB;
            font-size: 14px;
            margin-top: 4px;
        }
    </style>

    <div class="rme-header">
        <img src="https://raw.githubusercontent.com/Rohit113114/rme-quote-generator/main/rme_logo.png">
        <div>
            <div class="rme-header-title">RME Commercial Dashboard</div>
            <div class="rme-header-subtitle">Quotation • Workflow • PO • Invoice Tracking</div>
        </div>
    </div>
    """,
    unsafe_allow_html=True
)


CUSTOMERS_FILE = Path("customers.xlsx")
QUOTE_TEMPLATE_FILE = Path("rme_excel_template.xlsx")
INVOICE_TEMPLATE_FILE = Path("invoice_template.xlsx")

GOOGLE_SHEET_NAME = "RME Quote Register"
GOOGLE_DRIVE_QUOTE_FOLDER = "RME Generated Quotes"
MAX_ITEM_ROWS = 12
MAX_RETRIES = 5
MAXIMUM_BACKOFF = 32

REGISTER_HEADERS = [
    "Quote Number",
    "Revision",
    "Created Date",
    "Customer",
    "Department",
    "Company",
    "Job Status",
    "PO Number",
    "Invoice Number",
    "Quote Released Date",
    "PO Received Date",
    "Item Delivered Date",
    "Invoice Sent Date",
    "Invoice Due Date",
    "Invoice Paid Date",
    "Job Completed Date",
    "Subtotal",
    "GST",
    "Total",
    "Quote Excel Link",
    "Quote PDF Link",
]

STATUS_OPTIONS = [
    "Draft",
    "Released",
    "PO Received",
    "Items Delivered",
    "Invoice Sent",
    "Paid",
    "Completed",
    "Closed",
]

MONEY_COLUMNS = ["Subtotal", "GST", "Total"]


def is_blank(value):
    if value is None:
        return True

    try:
        if pd.isna(value):
            return True
    except Exception:
        pass

    return str(value).strip() in {"", "nan", "NaN", "NaT", "None"}


def clean_text(value):
    return "" if is_blank(value) else str(value).strip()


def generate_invoice_number(quote_number):
    return f"INV-{quote_number}"


def clean_money(value):
    if is_blank(value):
        return 0.0

    try:
        return float(str(value).replace("$", "").replace(",", "").strip())
    except Exception:
        return 0.0


def clean_bool(value, default=False):
    if is_blank(value):
        return default

    if isinstance(value, bool):
        return value

    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def parse_date(value):
    if is_blank(value):
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    for date_format in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(value).strip(), date_format).date()
        except ValueError:
            pass

    parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    return None if pd.isna(parsed) else parsed.date()


def format_date(date_value):
    parsed_date = parse_date(date_value)
    return "" if parsed_date is None else parsed_date.strftime("%d/%m/%Y")


def paragraph_safe(value):
    text = clean_text(value).replace("\r\n", "\n").replace("\r", "\n")
    return escape(text).replace("\n", "<br/>") or " "


def generate_hebrew_quote_prefix():
    today = dates.GregorianDate.today()
    hebrew_date = today.to_heb()

    month_codes = {
        1: "NS",
        2: "IY",
        3: "SV",
        4: "TM",
        5: "AV",
        6: "EL",
        7: "TS",
        8: "CH",
        9: "KS",
        10: "TV",
        11: "SH",
        12: "AD",
        13: "A2",
    }

    return f"{hebrew_date.day:02d}{month_codes[hebrew_date.month]}{hebrew_date.year}"


def generate_next_quote_number(register_df=None):
    prefix = generate_hebrew_quote_prefix()
    next_sequence = 1

    if register_df is not None and "Quote Number" in register_df.columns:
        pattern = re.compile(rf"^{re.escape(prefix)}(?:-(\d{{3}}))?$")

        for quote_number in register_df["Quote Number"].dropna().astype(str):
            match = pattern.match(quote_number.strip())

            if match:
                sequence = int(match.group(1) or "1")
                next_sequence = max(next_sequence, sequence + 1)

    if next_sequence == 1:
        return prefix

    return f"{prefix}-{next_sequence:03d}"


def run_with_google_retries(action_name, operation):
    last_exception = None

    for retry in range(MAX_RETRIES):
        try:
            return operation()
        except Exception as exc:
            last_exception = exc

            if retry == MAX_RETRIES - 1:
                break

            wait_time = min((2 ** retry) + random.uniform(0, 1), MAXIMUM_BACKOFF)
            st.warning(f"{action_name} failed temporarily. Retrying in {wait_time:.1f} seconds...")
            time.sleep(wait_time)

    raise last_exception


@st.cache_resource(show_spinner=False)
def get_google_client():
    if "gcp_service_account" not in st.secrets:
        raise RuntimeError("Google service account settings are missing from Streamlit Secrets.")

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]

    credentials = Credentials.from_service_account_info(
        st.secrets["gcp_service_account"],
        scopes=scopes,
    )

    return gspread.authorize(credentials)


@st.cache_resource(show_spinner=False)
def get_drive_service():
    if "gcp_service_account" not in st.secrets:
        raise RuntimeError("Google service account settings are missing from Streamlit Secrets.")

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]

    credentials = Credentials.from_service_account_info(
        st.secrets["gcp_service_account"],
        scopes=scopes,
    )

    return build("drive", "v3", credentials=credentials)


def make_drive_file_viewable(file_id):
    drive_service = get_drive_service()

    try:
        drive_service.permissions().create(
            fileId=file_id,
            body={"type": "anyone", "role": "reader"},
            fields="id",
        ).execute()
    except Exception:
        pass


def get_or_create_drive_folder(folder_name):
    drive_service = get_drive_service()

    safe_folder_name = folder_name.replace("'", "\\'")
    query = (
        f"name = '{safe_folder_name}' "
        "and mimeType = 'application/vnd.google-apps.folder' "
        "and trashed = false"
    )

    response = drive_service.files().list(
        q=query,
        spaces="drive",
        fields="files(id, name)",
    ).execute()

    folders = response.get("files", [])

    if folders:
        return folders[0]["id"]

    folder_metadata = {
        "name": folder_name,
        "mimeType": "application/vnd.google-apps.folder",
    }

    folder = drive_service.files().create(
        body=folder_metadata,
        fields="id",
    ).execute()

    make_drive_file_viewable(folder["id"])
    return folder["id"]


def upload_file_to_drive(file_bytes, filename, mime_type, folder_id):
    drive_service = get_drive_service()

    file_metadata = {
        "name": filename,
        "parents": [folder_id],
    }

    media = MediaIoBaseUpload(
        BytesIO(file_bytes),
        mimetype=mime_type,
        resumable=False,
    )

    uploaded_file = drive_service.files().create(
        body=file_metadata,
        media_body=media,
        fields="id, webViewLink",
    ).execute()

    make_drive_file_viewable(uploaded_file["id"])

    return uploaded_file.get("webViewLink", "")


def connect_google_sheet():
    sheet_name = st.secrets.get("google_sheet_name", GOOGLE_SHEET_NAME)

    return run_with_google_retries(
        "Google Sheets connection",
        lambda: get_google_client().open(sheet_name).sheet1,
    )


def update_sheet_range(sheet, range_name, values, value_input_option="USER_ENTERED"):
    return sheet.update(
        range_name=range_name,
        values=values,
        value_input_option=value_input_option,
    )


def get_sheet_values(sheet):
    return run_with_google_retries("Quote register read", sheet.get_all_values)


def ensure_register_headers(sheet):
    all_values = get_sheet_values(sheet)
    headers = [clean_text(header) for header in all_values[0]] if all_values else []

    if not headers or not any(headers):
        last_column = get_column_letter(len(REGISTER_HEADERS))
        update_sheet_range(sheet, f"A1:{last_column}1", [REGISTER_HEADERS], "RAW")
        return REGISTER_HEADERS.copy()

    missing_headers = [header for header in REGISTER_HEADERS if header not in headers]
    normalized_headers = headers + missing_headers

    if normalized_headers != headers:
        last_column = get_column_letter(len(normalized_headers))
        update_sheet_range(sheet, f"A1:{last_column}1", [normalized_headers], "RAW")

    return normalized_headers


@st.cache_data(ttl=30, show_spinner=False)
def get_register_dataframe():
    sheet = connect_google_sheet()
    headers = ensure_register_headers(sheet)
    all_values = get_sheet_values(sheet)

    records = []

    for row in all_values[1:]:
        padded_row = row + [""] * (len(headers) - len(row))
        record = dict(zip(headers, padded_row[:len(headers)]))

        if any(clean_text(value) for value in record.values()):
            records.append(record)

    register_df = pd.DataFrame(records) if records else pd.DataFrame(columns=headers)

    for header in REGISTER_HEADERS:
        if header not in register_df.columns:
            register_df[header] = ""

    ordered_columns = REGISTER_HEADERS + [
        column for column in register_df.columns if column not in REGISTER_HEADERS
    ]

    return register_df[ordered_columns]


def get_workbook():
    sheet_name = st.secrets.get("google_sheet_name", GOOGLE_SHEET_NAME)

    return run_with_google_retries(
        "Google workbook connection",
        lambda: get_google_client().open(sheet_name),
    )


def get_quote_items_sheet():
    workbook = get_workbook()

    try:
        return workbook.worksheet("Quote Items")
    except Exception:
        sheet = workbook.add_worksheet(title="Quote Items", rows=1000, cols=7)
        sheet.append_row(
            [
                "Quote Number",
                "Revision",
                "Part Number",
                "Description",
                "Qty",
                "Unit Price",
                "Line Total",
            ],
            value_input_option="USER_ENTERED",
        )
        return sheet


def save_quote_items(quote_number, revision, items):
    sheet = get_quote_items_sheet()

    for item in items:
        run_with_google_retries(
            "Quote items save",
            lambda item=item: sheet.append_row(
                [
                    clean_text(quote_number),
                    clean_text(revision),
                    item.get("part_no", ""),
                    item.get("description", ""),
                    item.get("qty", 0),
                    item.get("unit_price", 0),
                    item.get("total", 0),
                ],
                value_input_option="USER_ENTERED",
            ),
        )


def get_quote_items(quote_number):
    sheet = get_quote_items_sheet()
    records = sheet.get_all_records()

    matched_items = []
    selected_quote_number = str(quote_number).strip().replace(".0", "")

    for row in records:
        sheet_quote_number = str(row.get("Quote Number", "")).strip().replace(".0", "")

        if sheet_quote_number == selected_quote_number:
            matched_items.append(row)

    return matched_items


def clear_register_cache():
    try:
        get_register_dataframe.clear()
    except Exception:
        pass


def append_register_row(row_values):
    sheet = connect_google_sheet()
    headers = ensure_register_headers(sheet)

    row_by_header = dict(zip(REGISTER_HEADERS, row_values))
    aligned_row = [row_by_header.get(header, "") for header in headers]

    run_with_google_retries(
        "Quote register save",
        lambda: sheet.append_row(aligned_row, value_input_option="USER_ENTERED"),
    )

    clear_register_cache()


def update_register_row(quote_number, revision, update_values):
    sheet = connect_google_sheet()
    headers = ensure_register_headers(sheet)
    all_values = get_sheet_values(sheet)

    quote_col_index = headers.index("Quote Number")
    revision_col_index = headers.index("Revision")

    row_to_update = None
    row_values = None

    for row_number, row in enumerate(all_values[1:], start=2):
        padded_row = row + [""] * (len(headers) - len(row))

        if (
            clean_text(padded_row[quote_col_index]) == clean_text(quote_number)
            and clean_text(padded_row[revision_col_index]) == clean_text(revision)
        ):
            row_to_update = row_number
            row_values = padded_row[:len(headers)]
            break

    if row_to_update is None:
        return False

    for column_name, value in update_values.items():
        if column_name in headers:
            row_values[headers.index(column_name)] = value

    last_column = get_column_letter(len(headers))

    run_with_google_retries(
        "Quote register update",
        lambda: update_sheet_range(
            sheet,
            f"A{row_to_update}:{last_column}{row_to_update}",
            [row_values],
        ),
    )

    clear_register_cache()
    return True


def quote_revision_exists(register_df, quote_number, revision):
    if register_df.empty:
        return False

    quote_matches = register_df["Quote Number"].apply(clean_text) == clean_text(quote_number)
    revision_matches = register_df["Revision"].apply(clean_text) == clean_text(revision)

    return bool((quote_matches & revision_matches).any())


@st.cache_data(show_spinner=False)
def load_customers():
    if not CUSTOMERS_FILE.exists():
        raise FileNotFoundError(f"{CUSTOMERS_FILE} was not found.")

    customers = pd.read_excel(CUSTOMERS_FILE)
    customers.columns = customers.columns.astype(str).str.strip()

    required_columns = ["Contact Name", "Department", "Company", "Address", "City/State"]
    missing_columns = [column for column in required_columns if column not in customers.columns]

    if missing_columns:
        raise ValueError("customers.xlsx is missing required columns: " + ", ".join(missing_columns))

    if customers.empty:
        raise ValueError("customers.xlsx does not contain any customer rows.")

    return customers.fillna("")


def send_email_with_pdf(to_email, subject, body, pdf_bytes, pdf_filename):
    if "smtp" not in st.secrets:
        raise RuntimeError("SMTP email settings are not configured in Streamlit Secrets.")

    smtp_settings = st.secrets["smtp"]
    smtp_host = smtp_settings["host"]
    smtp_port = int(smtp_settings["port"])
    smtp_user = smtp_settings["username"]
    smtp_password = smtp_settings["password"]
    from_email = smtp_settings["from_email"]
    use_ssl = clean_bool(smtp_settings.get("use_ssl", smtp_port == 465), smtp_port == 465)

    msg = EmailMessage()
    msg["From"] = from_email
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.set_content(body)

    msg.add_attachment(
        pdf_bytes,
        maintype="application",
        subtype="pdf",
        filename=pdf_filename,
    )

    if use_ssl:
        with smtplib.SMTP_SSL(smtp_host, smtp_port) as smtp:
            smtp.login(smtp_user, smtp_password)
            smtp.send_message(msg)
    else:
        with smtplib.SMTP(smtp_host, smtp_port) as smtp:
            smtp.starttls()
            smtp.login(smtp_user, smtp_password)
            smtp.send_message(msg)


def format_quote_option(register_df, row_index):
    row = register_df.loc[row_index]
    parts = [f"{clean_text(row.get('Quote Number', ''))}-R{clean_text(row.get('Revision', ''))}"]

    customer = clean_text(row.get("Customer", ""))
    created_date = clean_text(row.get("Created Date", ""))

    if customer:
        parts.append(customer)

    if created_date:
        parts.append(created_date)

    return " | ".join(parts)


def create_excel_quote(
    quote_number,
    revision,
    selected_customer,
    department,
    company,
    address,
    city_state,
    scope,
    items,
    subtotal,
    gst,
    grand_total,
):
    if not QUOTE_TEMPLATE_FILE.exists():
        raise FileNotFoundError(f"{QUOTE_TEMPLATE_FILE} was not found.")

    if len(items) > MAX_ITEM_ROWS:
        raise ValueError(f"The Excel quote template supports up to {MAX_ITEM_ROWS} item rows.")

    workbook = load_workbook(QUOTE_TEMPLATE_FILE)
    worksheet = workbook.active

    worksheet["F8"] = quote_number
    worksheet["K8"] = revision
    worksheet["F9"] = datetime.today().strftime("%d/%m/%Y")

    worksheet["C13"] = selected_customer
    worksheet["C14"] = department
    worksheet["C15"] = company
    worksheet["C16"] = address
    worksheet["C17"] = city_state
    worksheet["B20"] = scope

    start_row = 26

    for row in range(start_row, start_row + MAX_ITEM_ROWS):
        for column in ("B", "C", "K", "L", "M"):
            worksheet[f"{column}{row}"] = None

    for index, item in enumerate(items):
        row = start_row + index
        worksheet[f"B{row}"] = item["part_no"]
        worksheet[f"C{row}"] = item["description"]
        worksheet[f"K{row}"] = item["qty"]
        worksheet[f"L{row}"] = item["unit_price"]
        worksheet[f"M{row}"] = f"=K{row}*L{row}"
        worksheet[f"L{row}"].number_format = "$#,##0.00"
        worksheet[f"M{row}"].number_format = "$#,##0.00"

    worksheet["L38"] = "=SUM(M26:M37)"
    worksheet["L39"] = "=L38*10%"
    worksheet["L40"] = "=L38+L39"

    for cell in ("L38", "L39", "L40"):
        worksheet[cell].number_format = "$#,##0.00"

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    return excel_buffer.getvalue()


def create_pdf_quote(
    quote_number,
    revision,
    selected_customer,
    department,
    company,
    address,
    city_state,
    scope,
    items,
    subtotal,
    gst,
    grand_total,
):
    pdf_buffer = BytesIO()

    document = SimpleDocTemplate(
        pdf_buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()
    table_cell_style = ParagraphStyle(
        "QuoteTableCell",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
    )

    elements = [
        Paragraph("Rail and Marine Engineering Pty Ltd", styles["Title"]),
        Paragraph("ACN 656374373 | ABN 82656374373 | Bibra Lake, Western Australia", styles["Normal"]),
        Spacer(1, 12),
        Paragraph(f"Quotation: {paragraph_safe(quote_number)} Rev {paragraph_safe(revision)}", styles["Heading2"]),
        Paragraph(f"Date: {datetime.today().strftime('%d/%m/%Y')}", styles["Normal"]),
        Spacer(1, 12),
    ]

    customer_text = f"""
    <b>Customer</b><br/>
    Name: {paragraph_safe(selected_customer)}<br/>
    Department: {paragraph_safe(department)}<br/>
    Company: {paragraph_safe(company)}<br/>
    Address: {paragraph_safe(address)}<br/>
    City/State: {paragraph_safe(city_state)}
    """

    elements += [
        Paragraph(customer_text, styles["Normal"]),
        Spacer(1, 12),
        Paragraph("<b>Description of work, scope and conditions</b>", styles["Normal"]),
        Paragraph(paragraph_safe(scope), styles["Normal"]),
        Spacer(1, 12),
    ]

    table_data = [["RME P/N", "Description", "Qty", "$ per unit", "$ Value"]]

    for item in items:
        table_data.append(
            [
                Paragraph(paragraph_safe(item["part_no"]), table_cell_style),
                Paragraph(paragraph_safe(item["description"]), table_cell_style),
                item["qty"],
                f"${item['unit_price']:,.2f}",
                f"${item['total']:,.2f}",
            ]
        )

    table_data.append(["", "", "", "Sub Total", f"${subtotal:,.2f}"])
    table_data.append(["", "", "", "10% GST", f"${gst:,.2f}"])
    table_data.append(["", "", "", "Total including GST", f"${grand_total:,.2f}"])

    table = Table(table_data, colWidths=[90, 300, 60, 100, 100], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.black),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("ALIGN", (2, 1), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (3, -3), (-1, -1), "Helvetica-Bold"),
            ]
        )
    )

    elements += [
        table,
        Spacer(1, 20),
        Paragraph("Contact Details", styles["Heading3"]),
        Paragraph("Rohit Saini | Mechanical Engineer | +610481247284 | rohit@rmerail.com", styles["Normal"]),
    ]

    document.build(elements)
    return pdf_buffer.getvalue()


try:
    customers_db = load_customers()
    customers_load_error = None
except Exception as exc:
    customers_db = pd.DataFrame()
    customers_load_error = exc


tab_dashboard, tab_create, tab_update, tab_invoice, tab_register = st.tabs(
    ["Dashboard", "Create New Quote", "Update Existing Quote", "Create Invoice", "Quote Register"]
)


with tab_dashboard:
    st.subheader("RME Quote Dashboard")

    try:
        dashboard_df = get_register_dataframe()

        if dashboard_df.empty:
            st.write("No quote data available.")
        else:
            for money_col in MONEY_COLUMNS:
                dashboard_df[money_col] = dashboard_df[money_col].apply(clean_money)

            dashboard_df["Parsed Due Date"] = dashboard_df["Invoice Due Date"].apply(parse_date)
            dashboard_df["Paid Blank"] = dashboard_df["Invoice Paid Date"].apply(is_blank)

            total_quotes = len(dashboard_df)
            paid_df = dashboard_df[
                dashboard_df["Job Status"].isin(["Paid", "Completed", "Closed"])
            ]
            
            total_revenue = paid_df["Total"].sum()
            paid_jobs = len(paid_df)

            po_received = len(dashboard_df[dashboard_df["Job Status"] == "PO Received"])
            invoice_sent = len(dashboard_df[dashboard_df["Job Status"] == "Invoice Sent"])

            overdue_df = dashboard_df[
                (dashboard_df["Parsed Due Date"].notna())
                & (dashboard_df["Parsed Due Date"] < date.today())
                & (dashboard_df["Paid Blank"])
            ]

            col1, col2, col3 = st.columns(3)
            col1.metric("Total Quotes", total_quotes)
            col2.metric("Gross Revenue", f"${total_revenue:,.2f}")
            col3.metric("Paid Jobs", paid_jobs)

            col4, col5, col6 = st.columns(3)
            col4.metric("PO Received", po_received)
            col5.metric("Invoices Sent", invoice_sent)
            col6.metric("Overdue Invoices", len(overdue_df))

            st.subheader("Quote Search")

            search_quote = st.text_input("Search Quote Number")
            search_customer = st.text_input("Search Customer")
            search_po = st.text_input("Search PO Number")
            search_status = st.selectbox("Filter Job Status", ["All"] + STATUS_OPTIONS)

            filtered_df = dashboard_df.copy()

            if search_quote:
                filtered_df = filtered_df[
                    filtered_df["Quote Number"].astype(str).str.contains(search_quote, case=False, na=False)
                ]

            if search_customer:
                filtered_df = filtered_df[
                    filtered_df["Customer"].astype(str).str.contains(search_customer, case=False, na=False)
                ]

            if search_po:
                filtered_df = filtered_df[
                    filtered_df["PO Number"].astype(str).str.contains(search_po, case=False, na=False)
                ]

            if search_status != "All":
                filtered_df = filtered_df[filtered_df["Job Status"].astype(str) == search_status]

            display_df = filtered_df.drop(columns=["Parsed Due Date", "Paid Blank"], errors="ignore").copy()

            for money_col in MONEY_COLUMNS:
                display_df[money_col] = display_df[money_col].apply(lambda value: f"${clean_money(value):,.2f}")

            st.subheader("Quote Results")
            st.dataframe(display_df, use_container_width=True)

            if len(overdue_df) > 0:
                overdue_display_df = overdue_df.drop(columns=["Parsed Due Date", "Paid Blank"], errors="ignore").copy()

                for money_col in MONEY_COLUMNS:
                    overdue_display_df[money_col] = overdue_display_df[money_col].apply(
                        lambda value: f"${clean_money(value):,.2f}"
                    )

                st.subheader("Overdue Invoices")
                st.dataframe(overdue_display_df, use_container_width=True)

    except Exception as exc:
        st.error("Dashboard failed to load.")
        st.error(exc)


with tab_create:
    st.subheader("Quote Details")

    if customers_load_error is not None:
        st.error("Customer database could not be loaded.")
        st.error(customers_load_error)
    else:
        try:
            register_for_numbering = get_register_dataframe()
            auto_quote_number = generate_next_quote_number(register_for_numbering)
        except Exception as exc:
            auto_quote_number = generate_next_quote_number()
            st.warning("Could not check Google Sheets for the next available quote number.")
            st.error(exc)

        quote_number = st.text_input("Quote Number", value=auto_quote_number)
        revision = st.text_input("Revision", "0")

        quote_reference = clean_text(quote_number)
        st.info(f"Quote Reference: {quote_reference}")

        st.subheader("Internal Workflow Tracking")

        job_status = st.selectbox("Job Status", STATUS_OPTIONS)
        po_number = st.text_input("PO Number")
        invoice_number = st.text_input("Invoice Number")

        quote_released_date = st.date_input("Date Quote Released", value=None)
        po_received_date = st.date_input("Date PO Received", value=None)
        item_delivered_date = st.date_input("Date Item Delivered", value=None)
        invoice_sent_date = st.date_input("Date Invoice Sent", value=None)
        invoice_due_date = st.date_input("Invoice Due Date", value=None)
        invoice_paid_date = st.date_input("Date Invoice Paid", value=None)
        job_completed_date = st.date_input("Date Job Completed", value=None)

        st.subheader("Customer Details")

        customer_names = customers_db["Contact Name"].astype(str).tolist()
        selected_customer = st.selectbox("Customer Contact", customer_names)

        customer_row = customers_db[
            customers_db["Contact Name"].astype(str) == selected_customer
        ].iloc[0]

        department = clean_text(customer_row["Department"])
        company = clean_text(customer_row["Company"])
        address = clean_text(customer_row["Address"])
        city_state = clean_text(customer_row["City/State"])
        customer_email = clean_text(customer_row["Email"]) if "Email" in customers_db.columns else ""

        st.write(f"Name: {selected_customer}")
        st.write(f"Department: {department}")
        st.write(f"Company: {company}")
        st.write(f"Address: {address}")
        st.write(f"City/State: {city_state}")

        if customer_email:
            st.write(f"Email: {customer_email}")

        scope = st.text_area("Scope of Work")

        st.subheader("Items")

        item_count = st.number_input(
            "Number of item rows",
            min_value=1,
            max_value=MAX_ITEM_ROWS,
            value=3,
        )

        items = []

        for i in range(item_count):
            st.markdown(f"### Item {i + 1}")

            part_no = st.text_input(f"Part Number {i + 1}", key=f"part{i}")
            description = st.text_input(f"Description {i + 1}", key=f"desc{i}")

            qty = st.number_input(f"Qty {i + 1}", min_value=0, value=0, key=f"qty{i}")
            unit_price = st.number_input(f"Unit Price {i + 1}", min_value=0.0, value=0.0, key=f"price{i}")

            if qty > 0:
                items.append(
                    {
                        "part_no": clean_text(part_no),
                        "description": clean_text(description),
                        "qty": qty,
                        "unit_price": unit_price,
                        "total": qty * unit_price,
                    }
                )

        subtotal = sum(item["total"] for item in items)
        gst = subtotal * 0.10
        grand_total = subtotal + gst

        st.subheader("Totals")
        st.write(f"Subtotal: ${subtotal:,.2f}")
        st.write(f"GST: ${gst:,.2f}")
        st.write(f"Grand Total: ${grand_total:,.2f}")

        if st.button("Generate Quote"):
            validation_errors = []

            if not clean_text(quote_number):
                validation_errors.append("Quote Number is required.")

            if not clean_text(revision):
                validation_errors.append("Revision is required.")

            if not items:
                validation_errors.append("Add at least one item with a quantity greater than zero.")

            duplicate_found = False

            if not validation_errors:
                try:
                    duplicate_found = quote_revision_exists(
                        get_register_dataframe(),
                        quote_number,
                        revision,
                    )
                except Exception as exc:
                    st.warning("Could not check for duplicate quote numbers before generating.")
                    st.error(exc)

            if duplicate_found:
                validation_errors.append(f"Quote {quote_reference} already exists in the register.")

            if validation_errors:
                for error in validation_errors:
                    st.error(error)
            else:
                try:
                    excel_bytes = create_excel_quote(
                        quote_number,
                        revision,
                        selected_customer,
                        department,
                        company,
                        address,
                        city_state,
                        scope,
                        items,
                        subtotal,
                        gst,
                        grand_total,
                    )

                    pdf_bytes = create_pdf_quote(
                        quote_number,
                        revision,
                        selected_customer,
                        department,
                        company,
                        address,
                        city_state,
                        scope,
                        items,
                        subtotal,
                        gst,
                        grand_total,
                    )

                    excel_filename = f"RME_Quote_{quote_reference}.xlsx"
                    pdf_filename = f"RME_Quote_{quote_reference}.pdf"
                    excel_drive_link = ""
                    pdf_drive_link = ""

                    try:
                        quote_folder_id = get_or_create_drive_folder(GOOGLE_DRIVE_QUOTE_FOLDER)

                        excel_drive_link = upload_file_to_drive(
                            excel_bytes,
                            excel_filename,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            quote_folder_id,
                        )

                        pdf_drive_link = upload_file_to_drive(
                            pdf_bytes,
                            pdf_filename,
                            "application/pdf",
                            quote_folder_id,
                        )
                    except Exception as exc:
                        st.warning("Quote generated, but Google Drive upload failed.")
                        st.error(exc)

                    history_row = [
                        clean_text(quote_number),
                        clean_text(revision),
                        datetime.today().strftime("%d/%m/%Y"),
                        selected_customer,
                        department,
                        company,
                        job_status,
                        clean_text(po_number),
                        clean_text(invoice_number),
                        format_date(quote_released_date),
                        format_date(po_received_date),
                        format_date(item_delivered_date),
                        format_date(invoice_sent_date),
                        format_date(invoice_due_date),
                        format_date(invoice_paid_date),
                        format_date(job_completed_date),
                        subtotal,
                        gst,
                        grand_total,
                        excel_drive_link,
                        pdf_drive_link,
                    ]

                    try:
                        append_register_row(history_row)
                        save_quote_items(quote_number, revision, items)
                        st.success("Quote generated and saved to Google Sheets.")
                    except Exception as exc:
                        st.warning("Quote generated, but Google Sheets save failed.")
                        st.error(exc)

                    generation_token = datetime.now().strftime("%Y%m%d%H%M%S%f")

                    st.session_state["generated_quote"] = {
                        "quote_reference": quote_reference,
                        "excel_bytes": excel_bytes,
                        "pdf_bytes": pdf_bytes,
                        "excel_filename": f"RME_Quote_{quote_reference}.xlsx",
                        "pdf_filename": f"RME_Quote_{quote_reference}.pdf",
                        "customer_email": customer_email,
                        "generation_token": generation_token,
                    }

                except Exception as exc:
                    st.error("Quote generation failed.")
                    st.error(exc)

        generated_quote = st.session_state.get("generated_quote")

        if generated_quote:
            st.subheader("Generated Quote")

            st.download_button(
                label="Download Quote Excel",
                data=generated_quote["excel_bytes"],
                file_name=generated_quote["excel_filename"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"excel_download_{generated_quote['generation_token']}",
            )

            st.download_button(
                label="Download Quote PDF",
                data=generated_quote["pdf_bytes"],
                file_name=generated_quote["pdf_filename"],
                mime="application/pdf",
                key=f"pdf_download_{generated_quote['generation_token']}",
            )

            if generated_quote.get("excel_drive_link"):
                st.markdown(f"[Open Quote Excel in Google Drive]({generated_quote['excel_drive_link']})")

            if generated_quote.get("pdf_drive_link"):
                st.markdown(f"[Open Quote PDF in Google Drive]({generated_quote['pdf_drive_link']})")

            st.subheader("Email Quote")

            with st.form(f"email_quote_form_{generated_quote['generation_token']}"):
                email_to = st.text_input("Send to email", value=generated_quote["customer_email"])
                email_subject = st.text_input(
                    "Email subject",
                    value=f"RME Quotation {generated_quote['quote_reference']}",
                )
                email_body = st.text_area(
                    "Email body",
                    value=f"""Hi,

Please find attached RME quotation {generated_quote['quote_reference']}.

Regards,
Rohit Saini
Rail and Marine Engineering Pty Ltd""",
                )
                send_email = st.form_submit_button("Send Quote Email")

            if send_email:
                if not clean_text(email_to):
                    st.error("Recipient email address is required.")
                else:
                    try:
                        send_email_with_pdf(
                            clean_text(email_to),
                            email_subject,
                            email_body,
                            generated_quote["pdf_bytes"],
                            generated_quote["pdf_filename"],
                        )
                        st.success("Email sent successfully.")
                    except Exception as exc:
                        st.error("Email sending failed.")
                        st.error(exc)


with tab_update:
    st.subheader("Update Existing Quote")

    try:
        register_df = get_register_dataframe()

        if register_df.empty:
            st.write("No quote records found.")
        else:
            selected_row_index = st.selectbox(
                "Select Quote",
                register_df.index.tolist(),
                format_func=lambda row_index: format_quote_option(register_df, row_index),
            )

            selected_record = register_df.loc[selected_row_index]

            selected_quote_number = clean_text(selected_record.get("Quote Number", ""))
            selected_revision = clean_text(selected_record.get("Revision", ""))

            st.write("Customer:", clean_text(selected_record.get("Customer", "")))
            st.write("Company:", clean_text(selected_record.get("Company", "")))
            st.write("Current Status:", clean_text(selected_record.get("Job Status", "")))

            current_status = clean_text(selected_record.get("Job Status", "Draft"))

            if current_status not in STATUS_OPTIONS:
                current_status = "Draft"

            updated_job_status = st.selectbox(
                "Update Job Status",
                STATUS_OPTIONS,
                index=STATUS_OPTIONS.index(current_status),
            )

            updated_po_number = st.text_input(
                "Update PO Number",
                value=clean_text(selected_record.get("PO Number", "")),
            )

            updated_invoice_number = st.text_input(
                "Update Invoice Number",
                value=clean_text(selected_record.get("Invoice Number", "")),
            )

            update_quote_released_date = st.text_input(
                "Quote Released Date",
                value=clean_text(selected_record.get("Quote Released Date", "")),
            )

            update_po_received_date = st.text_input(
                "PO Received Date",
                value=clean_text(selected_record.get("PO Received Date", "")),
            )

            update_item_delivered_date = st.text_input(
                "Item Delivered Date",
                value=clean_text(selected_record.get("Item Delivered Date", "")),
            )

            update_invoice_sent_date = st.text_input(
                "Invoice Sent Date",
                value=clean_text(selected_record.get("Invoice Sent Date", "")),
            )

            update_invoice_due_date = st.text_input(
                "Invoice Due Date",
                value=clean_text(selected_record.get("Invoice Due Date", "")),
            )

            update_invoice_paid_date = st.text_input(
                "Invoice Paid Date",
                value=clean_text(selected_record.get("Invoice Paid Date", "")),
            )

            update_job_completed_date = st.text_input(
                "Job Completed Date",
                value=clean_text(selected_record.get("Job Completed Date", "")),
            )

            if st.button("Update Quote Register"):
                today_string = datetime.today().strftime("%d/%m/%Y")

                update_values = {
                    "Job Status": updated_job_status,
                    "PO Number": clean_text(updated_po_number),
                    "Invoice Number": clean_text(updated_invoice_number),
                    "Quote Released Date": clean_text(update_quote_released_date),
                    "PO Received Date": clean_text(update_po_received_date),
                    "Item Delivered Date": clean_text(update_item_delivered_date),
                    "Invoice Sent Date": clean_text(update_invoice_sent_date),
                    "Invoice Due Date": clean_text(update_invoice_due_date),
                    "Invoice Paid Date": clean_text(update_invoice_paid_date),
                    "Job Completed Date": clean_text(update_job_completed_date),
                }

                if updated_job_status == "Released" and not update_values["Quote Released Date"]:
                    update_values["Quote Released Date"] = today_string

                if updated_job_status == "PO Received" and not update_values["PO Received Date"]:
                    update_values["PO Received Date"] = today_string

                if updated_job_status == "Items Delivered" and not update_values["Item Delivered Date"]:
                    update_values["Item Delivered Date"] = today_string

                if updated_job_status == "Invoice Sent" and not update_values["Invoice Sent Date"]:
                    update_values["Invoice Sent Date"] = today_string

                if updated_job_status == "Paid" and not update_values["Invoice Paid Date"]:
                    update_values["Invoice Paid Date"] = today_string

                if updated_job_status == "Completed" and not update_values["Job Completed Date"]:
                    update_values["Job Completed Date"] = today_string

                success = update_register_row(
                    selected_quote_number,
                    selected_revision,
                    update_values,
                )

                if success:
                    st.success("Quote register updated successfully.")
                else:
                    st.error("Quote number/revision not found in register.")

    except Exception as exc:
        st.error("Could not load quote register.")
        st.error(exc)


with tab_invoice:
    st.subheader("Create Invoice")

    try:
        register_df = get_register_dataframe()

        if register_df.empty:
            st.write("No quote records found.")
        else:
            selected_invoice_quote = st.selectbox(
                "Select Quote to Invoice",
                register_df["Quote Number"].astype(str).tolist(),
                key="invoice_quote_select",
            )

            selected_invoice_record = register_df[
                register_df["Quote Number"].astype(str) == selected_invoice_quote
            ].iloc[0]

            invoice_quote_number = clean_text(selected_invoice_record.get("Quote Number", ""))
            invoice_number = generate_invoice_number(invoice_quote_number)

            st.info(f"Invoice Number: {invoice_number}")

            st.subheader("Customer Details")

            invoice_customer = clean_text(selected_invoice_record.get("Customer", ""))
            invoice_department = clean_text(selected_invoice_record.get("Department", ""))
            invoice_company = clean_text(selected_invoice_record.get("Company", ""))

            customer_match = customers_db[
                customers_db["Contact Name"].astype(str).str.strip().str.lower()
                == invoice_customer.strip().lower()
            ]

            if not customer_match.empty:
                customer_invoice_row = customer_match.iloc[0]
                invoice_address = clean_text(customer_invoice_row.get("Address", ""))
                invoice_city_state = clean_text(customer_invoice_row.get("City/State", ""))
            else:
                invoice_address = ""
                invoice_city_state = ""

            st.write(f"Customer: {invoice_customer}")
            st.write(f"Department: {invoice_department}")
            st.write(f"Company: {invoice_company}")
            st.write(f"Address: {invoice_address}")
            st.write(f"City / State: {invoice_city_state}")

            st.subheader("Billing Address")

            billing_address = st.text_area(
                "Billing Address",
                value="""The Pilbara Infrastructure Pty Ltd
Level 8, 256 St Georges Terrace
PERTH
WA 6000 Australia
invoices@fortescue.com""",
            )

            po_number_for_invoice = st.text_input(
                "Purchase Order Number",
                value=clean_text(selected_invoice_record.get("PO Number", "")),
            )

            invoice_items = get_quote_items(invoice_quote_number)

            st.subheader("Invoice Items")

            if invoice_items:
                st.dataframe(pd.DataFrame(invoice_items), use_container_width=True)
            else:
                st.warning("No saved item lines found for this quote. A summary invoice line will be used.")

            def create_invoice_from_quote():
                if not INVOICE_TEMPLATE_FILE.exists():
                    raise FileNotFoundError(f"{INVOICE_TEMPLATE_FILE} was not found.")

                wb = load_workbook(INVOICE_TEMPLATE_FILE)
                ws = wb.active

                if invoice_items:
                    subtotal = sum(clean_money(item.get("Line Total", 0)) for item in invoice_items)
                else:
                    subtotal = clean_money(selected_invoice_record.get("Subtotal", 0))

                gst = subtotal * 0.10
                total = subtotal + gst

                ws["F10"] = invoice_number
                ws["F11"] = datetime.today().strftime("%d/%m/%Y")
                ws["F12"] = po_number_for_invoice

                ws["B17"] = f"{invoice_customer} - {invoice_department}"
                ws["D18"] = invoice_company
                ws["D19"] = invoice_address
                ws["D20"] = invoice_city_state

                ws["I17"] = billing_address

                start_row = 27

                if invoice_items:
                    if len(invoice_items) > 9:
                        raise ValueError("Invoice template supports only 9 item rows before totals.")

                    for index, item in enumerate(invoice_items):
                        row = start_row + index

                        part_number = ""
                        description = ""

                        for key, value in item.items():
                            clean_key = str(key).strip()

                            if clean_key == "Part Number":
                                part_number = str(value).strip()

                            if clean_key == "Description":
                                description = str(value).strip()

                        invoice_description = f"{part_number} - {description}".strip(" -")

                        ws[f"B{row}"] = invoice_description
                        ws[f"I{row}"] = clean_money(item.get("Qty", 0))
                        ws[f"J{row}"] = clean_money(item.get("Unit Price", 0))
                        ws[f"K{row}"] = f"=I{row}*J{row}"

                        ws[f"J{row}"].number_format = "$#,##0.00"
                        ws[f"K{row}"].number_format = "$#,##0.00"

                        ws[f"B{row}"].alignment = Alignment(horizontal="left")
                        ws[f"I{row}"].alignment = Alignment(horizontal="center")
                        ws[f"J{row}"].alignment = Alignment(horizontal="right")
                        ws[f"K{row}"].alignment = Alignment(horizontal="right")
                else:
                    ws["B27"] = f"Quotation {invoice_quote_number}"
                    ws["I27"] = 1
                    ws["J27"] = subtotal
                    ws["K27"] = "=I27*J27"
                    ws["J27"].number_format = "$#,##0.00"
                    ws["K27"].number_format = "$#,##0.00"
                    ws["B27"].alignment = Alignment(horizontal="left")
                    ws["I27"].alignment = Alignment(horizontal="center")
                    ws["J27"].alignment = Alignment(horizontal="right")
                    ws["K27"].alignment = Alignment(horizontal="right")

                ws["K36"] = "=SUM(K27:K35)"
                ws["K37"] = "=K36*10%"
                ws["K38"] = "=K36+K37"

                for cell in ["K36", "K37", "K38"]:
                    ws[cell].number_format = "$#,##0.00"
                    ws[cell].alignment = Alignment(horizontal="right")

                wb.calculation.fullCalcOnLoad = True
                wb.calculation.forceFullCalc = True

                excel_buffer = BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)

                return excel_buffer.getvalue()

            if st.button("Generate Invoice Excel", key="generate_invoice_excel_button"):
                invoice_file = create_invoice_from_quote()

                st.success(f"Invoice generated: {invoice_number}")

                st.download_button(
                    label="Download Invoice Excel",
                    data=invoice_file,
                    file_name=f"RME_Invoice_{invoice_number}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

    except Exception as exc:
        st.error("Could not generate invoice.")
        st.error(exc)


with tab_register:
    st.subheader("Quote Register")

    try:
        register_df = get_register_dataframe()

        if not register_df.empty:
            display_register_df = register_df.copy()

            for money_col in MONEY_COLUMNS:
                display_register_df[money_col] = display_register_df[money_col].apply(
                    lambda value: f"${clean_money(value):,.2f}"
                )

            st.dataframe(display_register_df, use_container_width=True)

            csv_data = register_df.to_csv(index=False).encode("utf-8")

            st.download_button(
                label="Download Quote Register",
                data=csv_data,
                file_name="rme_quote_register.csv",
                mime="text/csv",
            )
        else:
            st.write("No quote records found yet.")

    except Exception as exc:
        st.write("Quote register will appear here after Google Sheets connection is active.")
        st.error(exc)
